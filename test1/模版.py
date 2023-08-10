import pandas as pd
import openpyxl as op
from io import StringIO
from openpyxl import Workbook


df = pd.DataFrame()                                    #标题栏
df.to_excel(r'C:\Users\admin\Desktop\test\test1\模版.xlsx')                                 #新表格路径
print("New excel had been created.")

data = pd.read_excel(io=r'C:\Users\admin\Desktop\test\test1\盈合公司信息平台.xlsx')  #读数据
wb = op.load_workbook(r'C:\Users\admin\Desktop\test\test1\模版.xlsx')   #选中目标excel  
ws = wb.worksheets[0]                                                  #选中目标sheet，从0开始


for row1 in range(0,20):                                               #一共写几行
    title_new = '文件《'+data['title'].iloc[row1]+'》的内容是？'         #生成问题
    ws.cell(row=row1+1, column=1).value = title_new                    #写入问题
    ws.cell(row=row1+1, column=2).value = data['answer'].iloc[row1]    #写入答案

wb.save("模版.xlsx")                 #保存